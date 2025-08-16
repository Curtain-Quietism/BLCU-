import os
from collections import defaultdict
import sys
import pandas as pd
import matplotlib
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator
from matplotlib import font_manager as fm
from matplotlib.axes import Axes
from matplotlib.patches import Rectangle
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from typing import Dict, List, Any, Tuple, cast

# ==============================================================================
# Configuration & Constants
# ==============================================================================

# --- File Paths ---
# Note: Per your request, these paths remain unchanged.
OUTPUT_IMAGE_DIR = "img"
OUTPUT_EXCEL_PATH = "output/信息统计汇总.xlsx"

# --- Charting Parameters ---
TOP_N_PROVINCES = 15

# --- Excel Styling ---
THIN_BORDER_SIDE = Side(border_style="thin", color="000000")
TABLE_BORDER = Border(
    left=THIN_BORDER_SIDE,
    right=THIN_BORDER_SIDE,
    top=THIN_BORDER_SIDE,
    bottom=THIN_BORDER_SIDE,
)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
HEADER_FONT = Font(bold=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=False)

# --- Plotting Style & Palette ---
PALETTE = [
    "#4C78A8",
    "#F58518",
    "#E45756",
    "#72B7B2",
    "#54A24B",
    "#EECA3B",
    "#B279A2",
    "#FF9DA6",
    "#9C755F",
    "#BAB0AC",
    "#2E7EAF",
    "#FFB55A",
]

# ==============================================================================
# Utility Functions
# ==============================================================================


def ensure_directory_exists(dir_path: str) -> None:
    """Creates a directory if it does not already exist."""
    if dir_path and not os.path.exists(dir_path):
        os.makedirs(dir_path, exist_ok=True)


def format_as_integer_string(n: Any) -> str:
    """Formats a number as a comma-separated integer string."""
    try:
        return f"{int(n):,}"
    except (ValueError, TypeError):
        return str(n)


# ==============================================================================
# Data Processing Class
# ==============================================================================


class DataProcessor:
    """Handles reading and analyzing data from an Excel file."""

    def __init__(self, required_columns: List[str]):
        self.required_columns = required_columns

    def find_and_read_excel(self, directory: str) -> pd.DataFrame:
        """
        Finds and reads the first .xlsx or .xls file in the specified directory.

        Args:
            directory: The directory to search for an Excel file.

        Returns:
            A pandas DataFrame containing the Excel data.

        Raises:
            FileNotFoundError: If no Excel file is found in the directory.
        """
        for filename in os.listdir(directory):
            if filename.endswith((".xlsx", ".xls")):
                filepath = os.path.join(directory, filename)
                try:
                    print(f"Reading Excel file: {filename}")
                    return pd.read_excel(filepath, sheet_name=0)
                except Exception as e:
                    print(f"Error reading file {filename}: {e}")
        raise FileNotFoundError(f"No Excel file (.xlsx, .xls) found in '{directory}'.")

    def analyze(self, dataframe: pd.DataFrame) -> Dict[str, Any]:
        """
        Processes the DataFrame to generate summary statistics.

        Args:
            dataframe: The input DataFrame.

        Returns:
            A dictionary containing the analysis results.

        Raises:
            ValueError: If the DataFrame is missing required columns.
        """
        self._validate_columns(dataframe)

        # Process categorical data
        political_status = dataframe["政治面貌"].fillna("群众")
        analysis_results = {
            "录取类别": dataframe["录取类别"].value_counts(dropna=False).to_dict(),
            "省市": dataframe["省市"].value_counts(dropna=False).to_dict(),
            "性别": dataframe["性别"].value_counts(dropna=False).to_dict(),
            "民族": dataframe["民族"].value_counts(dropna=False).to_dict(),
            "政治面貌": political_status.value_counts(dropna=False).to_dict(),
            "出生日期": self._process_birthdays(dataframe["出生年月"]),
        }
        return analysis_results

    def _validate_columns(self, dataframe: pd.DataFrame) -> None:
        """Checks if all required columns are present in the DataFrame."""
        missing_columns = [
            col for col in self.required_columns if col not in dataframe.columns
        ]
        if missing_columns:
            raise ValueError(f"Data is missing required columns: {missing_columns}")

    def _process_birthdays(self, birthday_series: pd.Series) -> List[Dict[str, Any]]:
        """
        Processes the birthday column to get counts by year and month.

        This method is a more robust, pandas-idiomatic replacement for the original loop.
        """
        # Convert to datetime, coercing errors to NaT (Not a Time)
        birthdays = pd.to_datetime(birthday_series, errors="coerce").dropna()

        if birthdays.empty:
            return []

        # Group by year and month and count occurrences
        birthday_counts = birthdays.groupby(
            [birthdays.dt.year, birthdays.dt.month]
        ).size()

        # Restructure data into the nested format required by downstream functions
        year_month_map = defaultdict(list)
        for (year, month), count in birthday_counts.items():
            year_label = f"{year}年"
            month_label = f"{month}月"
            year_month_map[year_label].append({month_label: count})

        # Format into the final list of dictionaries
        final_result = [{year: months} for year, months in year_month_map.items()]
        return final_result


# ==============================================================================
# Chart Generation Class
# ==============================================================================


class ChartGenerator:
    """Generates and saves all charts based on analysis results."""

    def __init__(self, output_dir: str):
        self.output_dir = output_dir
        self._setup_matplotlib_style()
        ensure_directory_exists(self.output_dir)
        print(f"Chart output directory set to: '{self.output_dir}'")

    def _setup_matplotlib_style(self) -> None:
        """Configures Matplotlib settings for fonts and a clean theme."""
        # 1. Set Chinese Font
        font_candidates = [
            "Microsoft YaHei",
            "PingFang SC",
            "SimHei",
            "Noto Sans CJK SC",
            "Source Han Sans SC",
            "WenQuanYi Micro Hei",
        ]
        available_fonts = {f.name for f in fm.fontManager.ttflist}
        for name in font_candidates:
            if name in available_fonts:
                plt.rcParams["font.family"] = name
                break
        plt.rcParams["axes.unicode_minus"] = False

        # 2. Apply Clean Theme
        plt.rcParams.update(
            {
                "figure.dpi": 160,
                "savefig.dpi": 200,
                "axes.facecolor": "#FFFFFF",
                "figure.facecolor": "#FFFFFF",
                "axes.edgecolor": "#CCCCCC",
                "axes.titleweight": "bold",
                "axes.titlepad": 14,
                "axes.labelpad": 8,
                "axes.titlesize": 14,
                "axes.labelsize": 12,
                "xtick.labelsize": 11,
                "ytick.labelsize": 11,
                "grid.color": "#E6E6E6",
                "grid.linestyle": "-",
                "grid.linewidth": 1.0,
                "axes.grid": True,
                "axes.axisbelow": True,
                "legend.frameon": False,
            }
        )
        matplotlib.use("Agg")

    def _beautify_axes(self, ax: Axes) -> None:
        """Styles axes by removing top/right spines for a cleaner look."""
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.spines["left"].set_color("#CCCCCC")
        ax.spines["bottom"].set_color("#CCCCCC")

    def _add_bar_labels(
        self,
        ax: Axes,
        orientation: str = "v",
        show_pct: bool = False,
        total: int = None,
    ) -> None:
        """Annotates bars in a bar chart with their values and optional percentages."""
        offset = 3
        for patch in ax.patches:
            # Cast the generic Patch to a Rectangle to satisfy the type checker
            rect = cast(Rectangle, patch)
            if orientation == "v":
                value = rect.get_height()
                x = rect.get_x() + rect.get_width() / 2
                y = value
                xytext = (0, offset)
                ha, va = "center", "bottom"
            else:  # orientation == 'h'
                value = rect.get_width()
                x = value
                y = rect.get_y() + rect.get_height() / 2
                xytext = (offset + 3, 0)
                ha, va = "left", "center"

            label = f"{int(value):,}"
            if show_pct and total and total > 0:
                label += f" ({value / total:.1%})"

            ax.annotate(
                label,
                xy=(x, y),
                xytext=xytext,
                textcoords="offset points",
                ha=ha,
                va=va,
                fontsize=10,
                color="#333333",
            )

    def _save_figure(self, fig: plt.Figure, filename: str) -> None:
        """Saves a matplotlib figure to the output directory."""
        path = os.path.join(self.output_dir, filename)
        fig.tight_layout()
        fig.savefig(path, bbox_inches="tight")
        plt.close(fig)
        print(f"[Chart Saved] {path}")

    def generate_all(self, analysis_results: Dict[str, Any]) -> None:
        """
        Generates and saves all supported charts from the analysis results.
        """
        print("\n--- Starting Chart Generation ---")
        if isinstance(analysis_results.get("省市"), dict):
            self.plot_province_distribution(
                analysis_results["省市"]
            )
        if isinstance(analysis_results.get("性别"), dict):
            self.plot_gender_distribution(analysis_results["性别"])
        if isinstance(analysis_results.get("录取类别"), dict):
            self.plot_bar_chart(
                "录取类别分布", analysis_results["录取类别"], "录取类别_柱状.png"
            )
        if isinstance(analysis_results.get("民族"), dict):
            self.plot_ethnicity_distribution(analysis_results["民族"])
        if isinstance(analysis_results.get("政治面貌"), dict):
            self.plot_political_status_distribution(analysis_results["政治面貌"])
        if isinstance(analysis_results.get("出生日期"), list):
            self.plot_birth_year_distribution(analysis_results["出生日期"])
            self.plot_birth_month_distribution(analysis_results["出生日期"])
        print("--- Chart Generation Complete ---")

    def plot_province_distribution(self, counts: Dict[str, int]) -> None:
        """Generates a horizontal bar chart for province distribution."""
        if not counts:
            return

        items = sorted(counts.items(), key=lambda x: (-x[1], str(x[0])))
        labels = [item[0] for item in items][::-1]
        values = [item[1] for item in items][::-1]
        total = sum(counts.values())

        fig, ax = plt.subplots(figsize=(8, 6))
        ax.barh(labels, values, color=PALETTE)
        self._beautify_axes(ax)
        ax.set_title("省市分布")
        ax.set_xlabel("人数")
        ax.set_ylabel("省市")
        ax.xaxis.set_major_locator(MaxNLocator(integer=True))
        self._add_bar_labels(ax, orientation="h", show_pct=True, total=total)
        self._save_figure(fig, "省市分布.png")

    def plot_gender_distribution(self, counts: Dict[str, int]) -> None:
        """Generates a donut chart for gender distribution."""
        if not counts:
            return

        items = sorted(counts.items(), key=lambda x: str(x[0]))
        labels = [item[0] for item in items]
        values = [item[1] for item in items]
        total = sum(values)

        fig, ax = plt.subplots(figsize=(5.5, 5.5))
        result = ax.pie(
            values,
            startangle=90,
            counterclock=False,
            colors=PALETTE,
            wedgeprops={"width": 0.38, "edgecolor": "white"},
        )
        wedges = result[0]
        ax.set(aspect="equal", title="性别比例")

        # Center text
        male_count = counts.get("男", 0)
        female_count = counts.get("女", 0)
        ax.text(
            0,
            0.05,
            f"N = {format_as_integer_string(total)}",
            ha="center",
            va="center",
            fontsize=11,
        )
        if total > 0:
            ax.text(
                0,
                -0.10,
                f"男 {male_count / total:.1%} / 女 {female_count / total:.1%}",
                ha="center",
                va="center",
                fontsize=11,
            )

        # Legend
        legend_labels = [f"{lbl} {v / total:.1%}" for lbl, v in zip(labels, values)]
        ax.legend(
            wedges,
            legend_labels,
            loc="lower center",
            bbox_to_anchor=(0.5, -0.15),
            ncol=len(labels),
        )

        self._save_figure(fig, "性别_甜甜圈.png")

    def plot_bar_chart(self, title: str, counts: Dict[str, int], filename: str) -> None:
        """Generates a generic vertical bar chart."""
        if not counts:
            return

        items = sorted(counts.items(), key=lambda x: (-x[1], str(x[0])))
        labels = [item[0] for item in items]
        values = [item[1] for item in items]
        total = sum(values)

        fig, ax = plt.subplots(figsize=(8, 5))
        ax.bar(labels, values, color=PALETTE)
        self._beautify_axes(ax)
        ax.set(title=title, ylabel="人数")
        ax.yaxis.set_major_locator(MaxNLocator(integer=True))
        ax.set_xticklabels(labels, rotation=0)
        self._add_bar_labels(ax, orientation="v", show_pct=True, total=total)
        self._save_figure(fig, filename)

    def plot_ethnicity_distribution(self, counts: Dict[str, int]) -> None:
        """Generates a donut chart for ethnicity distribution."""
        if not counts:
            return

        items = sorted(counts.items(), key=lambda x: (-x[1], str(x[0])))
        labels = [item[0] for item in items]
        values = [item[1] for item in items]
        total = sum(values)

        fig, ax = plt.subplots(figsize=(7, 7))
        result = ax.pie(
            values,
            startangle=90,
            counterclock=False,
            colors=PALETTE,
            wedgeprops={"width": 0.35, "edgecolor": "white"},
        )
        wedges = result[0]
        ax.set(aspect="equal", title="民族分布（全部）")

        legend_labels = [
            f"{lbl} {v}人 ({v / total:.1%})" for lbl, v in zip(labels, values)
        ]
        ax.legend(
            wedges,
            legend_labels,
            loc="lower center",
            bbox_to_anchor=(0.5, -0.15),
            ncol=2 if len(labels) > 6 else 1,
        )
        self._save_figure(fig, "民族_甜甜圈.png")

    def plot_political_status_distribution(self, counts: Dict[str, int]) -> None:
        """Generates a donut chart for political status distribution."""
        if not counts:
            return

        items = sorted(counts.items(), key=lambda x: -x[1])
        labels = [item[0] for item in items]
        values = [item[1] for item in items]
        total = sum(values)

        fig, ax = plt.subplots(figsize=(6.2, 6.2))
        result = ax.pie(
            values,
            startangle=90,
            counterclock=False,
            colors=PALETTE,
            wedgeprops={"width": 0.35, "edgecolor": "white"},
        )
        wedges = result[0]
        ax.set(aspect="equal", title="政治面貌分布")

        legend_labels = [f"{lbl} {v / total:.1%}" for lbl, v in zip(labels, values)]
        ax.legend(
            wedges,
            legend_labels,
            loc="lower center",
            bbox_to_anchor=(0.5, -0.15),
            ncol=2 if len(labels) > 6 else 1,
        )
        self._save_figure(fig, "政治面貌_甜甜圈.png")

    def _parse_birthday_data(self, birthday_list: List) -> Tuple[Dict, Dict]:
        """Parses the nested birthday list into year and month counts."""
        years_count = defaultdict(int)
        months_count = defaultdict(int)
        for year_map in birthday_list:
            for year_label, month_list in year_map.items():
                year_sum = 0
                for month_entry in month_list:
                    for month_label, count in month_entry.items():
                        try:
                            month_num = int(str(month_label).replace("月", ""))
                            months_count[month_num] += int(count)
                            year_sum += int(count)
                        except (ValueError, TypeError):
                            continue
                years_count[year_label] += year_sum
        return dict(years_count), dict(months_count)

    def plot_birth_year_distribution(self, birthday_list: List) -> None:
        """Generates a bar chart for birth year distribution."""
        if not birthday_list:
            return
        years_count, _ = self._parse_birthday_data(birthday_list)

        # Sort by year numerically, not lexicographically
        items = sorted(years_count.items(), key=lambda x: int(x[0].replace("年", "")))

        self.plot_bar_chart("出生年份分布", dict(items), "出生年份_柱状.png")

    def plot_birth_month_distribution(self, birthday_list: List) -> None:
        """Generates a bar chart for birth month distribution across all years."""
        if not birthday_list:
            return
        _, months_count = self._parse_birthday_data(birthday_list)

        labels = [f"{i}月" for i in range(1, 13)]
        values = [months_count.get(i, 0) for i in range(1, 13)]
        total = sum(values)

        fig, ax = plt.subplots(figsize=(8, 4.8))
        ax.bar(labels, values, color=PALETTE)
        self._beautify_axes(ax)
        ax.set(title="出生月份分布（跨年份汇总）", ylabel="人数")
        ax.yaxis.set_major_locator(MaxNLocator(integer=True))
        self._add_bar_labels(ax, orientation="v", show_pct=False, total=total)
        self._save_figure(fig, "出生月份_柱状.png")


# ==============================================================================
# Excel Exporting Class
# ==============================================================================


class ExcelExporter:
    """Handles writing analysis results to a formatted Excel file."""

    def __init__(self, output_path: str):
        self.output_path = output_path
        ensure_directory_exists(os.path.dirname(output_path))
        print(f"Excel output path set to: '{self.output_path}'")

    def _auto_fit_columns(self, ws, min_width=10, max_width=40) -> None:
        """Adjusts column widths based on content."""
        col_widths = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                # Estimate length (double for CJK characters)
                length = sum(2 if ord(char) > 127 else 1 for char in str(cell.value))
                col_widths[cell.column] = max(col_widths.get(cell.column, 0), length)

        for col_idx, length in col_widths.items():
            width = min(max(min_width, length + 2), max_width)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def _style_table(self, ws, header_row=1) -> None:
        """Applies consistent styling to a worksheet table."""
        header_cells = ws[header_row]
        for cell in header_cells:
            cell.font = HEADER_FONT
            cell.alignment = ALIGN_CENTER
            cell.fill = HEADER_FILL
            cell.border = TABLE_BORDER

        for row in ws.iter_rows(min_row=header_row + 1):
            for cell in row:
                cell.border = TABLE_BORDER
                cell.alignment = (
                    ALIGN_RIGHT if isinstance(cell.value, (int, float)) else ALIGN_LEFT
                )

    def _write_counts_sheet(
        self, wb: Workbook, title: str, counts: Dict, key_col: str
    ) -> None:
        """Writes a standard key-value-percentage sheet."""
        ws = wb.create_sheet(title=title)
        ws.freeze_panes = "A2"
        ws.append([key_col, "人数", "占比"])

        rows = sorted(counts.items(), key=lambda x: (-x[1], str(x[0])))
        total = sum(v for _, v in rows) or 1

        for key, value in rows:
            ws.append([key, value, value / total])

        # Apply number formats
        for r in range(2, len(rows) + 2):
            ws.cell(row=r, column=2).number_format = "#,##0"
            ws.cell(row=r, column=3).number_format = "0.00%"

        self._style_table(ws)
        self._auto_fit_columns(ws)

    def _write_birthday_sheet(self, wb: Workbook, birthday_list: List) -> None:
        """Writes the specially formatted birthday statistics sheet."""
        ws = wb.create_sheet(title="出生日期")
        ws.freeze_panes = "A2"
        ws.append(["出生年份", "月份", "人数"])

        # Group data by year
        by_year = defaultdict(list)
        for year_map in birthday_list:
            for year_label, month_list in year_map.items():
                for month_entry in month_list:
                    mlabel, count = list(month_entry.items())[0]
                    by_year[year_label].append((mlabel, count))

        current_row = 2
        sorted_years = sorted(by_year.keys(), key=lambda s: int(s.replace("年", "")))

        for year in sorted_years:
            months_data = sorted(
                by_year[year], key=lambda m: int(m[0].replace("月", ""))
            )
            start_row_for_year = current_row
            year_total = 0

            for month_label, count in months_data:
                ws.cell(row=current_row, column=1, value=year).alignment = ALIGN_CENTER
                ws.cell(row=current_row, column=2, value=month_label)
                ws.cell(row=current_row, column=3, value=count).number_format = "#,##0"
                year_total += count
                current_row += 1

            # Subtotal Row
            ws.cell(row=current_row, column=2, value="合计").font = Font(bold=True)
            ws.cell(row=current_row, column=3, value=year_total).number_format = "#,##0"
            ws.cell(row=current_row, column=3, value=year_total).font = Font(bold=True)
            current_row += 1

            # Merge year cells
            if len(months_data) > 1:
                ws.merge_cells(
                    start_row=start_row_for_year,
                    start_column=1,
                    end_row=current_row - 2,
                    end_column=1,
                )

        self._style_table(ws)
        self._auto_fit_columns(ws)

    def export(self, analysis_results: Dict[str, Any]) -> None:
        """
        Exports the full analysis results dictionary to an Excel file.
        """
        print("\n--- Starting Excel Export ---")
        wb = Workbook()
        if wb.active is not None:
            wb.remove(wb.active)  # Remove default sheet

        sheet_mapping = {
            "录取类别": ("录取类别", "类别"),
            "省市": ("省市", "省市"),
            "性别": ("性别", "性别"),
            "民族": ("民族", "民族"),
            "政治面貌": ("政治面貌", "政治面貌"),
        }

        for key, (sheet_title, key_col) in sheet_mapping.items():
            if key in analysis_results and isinstance(analysis_results[key], dict):
                self._write_counts_sheet(
                    wb, sheet_title, analysis_results[key], key_col
                )

        if "出生日期" in analysis_results and isinstance(
            analysis_results["出生日期"], list
        ):
            self._write_birthday_sheet(wb, analysis_results["出生日期"])

        wb.save(self.output_path)
        print(f"Successfully generated Excel file: {self.output_path}")
        print("--- Excel Export Complete ---")


# ==============================================================================
# Main Execution
# ==============================================================================


def main():
    """Main script logic."""
    print("Starting data analysis script...")

    if getattr(sys, "frozen", False):
        current_directory = os.path.dirname(sys.executable)
    else:
        current_directory = os.path.dirname(os.path.abspath(__file__))

    # Define required columns for validation
    required_cols = ["录取类别", "省市", "性别", "民族", "政治面貌", "出生年月"]

    try:
        # 1. Process Data
        # current_directory = os.path.dirname(os.path.abspath(__file__))
        processor = DataProcessor(required_columns=required_cols)
        source_dataframe = processor.find_and_read_excel(directory=current_directory)
        analysis_results = processor.analyze(source_dataframe)

        # 2. Export to Excel
        exporter = ExcelExporter(output_path=OUTPUT_EXCEL_PATH)
        exporter.export(analysis_results)

        # 3. Generate Charts
        charter = ChartGenerator(output_dir=OUTPUT_IMAGE_DIR)
        charter.generate_all(analysis_results)

    except (FileNotFoundError, ValueError) as e:
        print(f"\nAn error occurred: {e}")
        print("Script execution halted.")
    except Exception as e:
        print(f"\nAn unexpected error occurred: {e}")
        print("Script execution halted.")

    print("\nScript finished.")


if __name__ == "__main__":
    main()
